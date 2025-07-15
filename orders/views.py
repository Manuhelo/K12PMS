import csv
import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime
from django.shortcuts import render, redirect
from django.contrib import messages
from .models import StudentOrder, OrderSummary, ProcurementThreshold
from inventory.models import Warehouse, InventoryItem
from products.models import EducationalProduct
from django.db.models import Sum
from django.http import HttpResponse
from functools import reduce
from django.db.models import Q
from django.contrib.auth.decorators import login_required


def upload_student_orders(request):
    if request.method == 'POST' and request.FILES.get('csv_file'):
        file = request.FILES['csv_file']
        decoded_file = file.read().decode('utf-8').splitlines()
        reader = csv.DictReader(decoded_file)

        created = 0
        skipped = 0

        for row in reader:
            try:
                warehouse = Warehouse.objects.get(name=row['Branch Name'].strip())
                sku = EducationalProduct.objects.get(sku=row['SKU Code'].strip())

                obj, created_flag = StudentOrder.objects.get_or_create(
                    academic_year=row['Academic Year'].strip(),
                    warehouse=warehouse,
                    enrollment_code=row['Enrollment Code'].strip(),
                    sku=sku,
                    defaults={
                        'student_name': row['Student Name'].strip(),
                        'gender': row['Gender'].strip(),
                        'student_class': row['Class'].strip(),
                        'quantity': int(row['Quantity']),
                        'paid_date': datetime.strptime(row['Paid Date'], '%Y-%m-%d').date(),
                    }
                )
                if created_flag:
                    created += 1
                else:
                    skipped += 1
            except Exception as e:
                print(f"Error: {e}")
                continue

        messages.success(request, f"Upload complete: {created} created, {skipped} skipped.")
        return redirect('upload_student_orders')

    return render(request, 'orders/upload.html')


def download_student_order_sample(request):
    response = HttpResponse(
        content_type='text/csv',
        headers={'Content-Disposition': 'attachment; filename="student_order_sample.csv"'},
    )

    writer = csv.writer(response)
    # Write headers
    writer.writerow([
        'Academic Year', 'Branch Name', 'Enrollment Code', 'Student Name',
        'Gender', 'Class', 'SKU Code', 'Quantity', 'Paid Date'
    ])

    # Sample data rows
    writer.writerow(['2025-26', 'Hyderabad Branch', 'ENR001', 'John Doe', 'Male', 'Grade 1', '4000001234', '3', '2025-07-01'])
    writer.writerow(['2025-26', 'Hyderabad Branch', 'ENR002', 'Jane Smith', 'Female', 'Grade 2', '4000005678', '2', '2025-07-02'])

    return response

def upload_order_summary(request):
    if request.method == "POST" and request.FILES.get("file"):
        file = request.FILES["file"]
        file_name = file.name.lower()

        rows = []
        errors = []

        # Parse file
        try:
            if file_name.endswith(".csv"):
                decoded_file = file.read().decode('utf-8').splitlines()
                reader = csv.DictReader(decoded_file)
                for row in reader:
                    rows.append(row)

            elif file_name.endswith((".xls", ".xlsx")):
                wb = openpyxl.load_workbook(file)
                sheet = wb.active
                headers = [str(cell.value).strip().lower() for cell in sheet[1]]
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    row_dict = dict(zip(headers, row))
                    rows.append(row_dict)

            else:
                messages.error(request, "Unsupported file format. Use CSV or Excel.")
                return redirect("upload_order_summary")
        except Exception as e:
            messages.error(request, f"Error reading file: {e}")
            return redirect("upload_order_summary")

        uploaded_count = 0

        for row in rows:
            try:
                warehouse_name = row.get("branch name", "").strip()
                grade = row.get("grade", "").strip()
                sku_code = str(row.get("sku code", "")).strip().upper()
                quantity = int(str(row.get("quantity", "0")).replace(",", "").strip())

                warehouse = Warehouse.objects.get(name__iexact=warehouse_name)
                sku = EducationalProduct.objects.get(sku=sku_code)

                summary, created = OrderSummary.objects.get_or_create(
                    warehouse=warehouse,
                    grade=grade,
                    sku=sku,
                    defaults={'total_quantity': quantity}
                )
                if not created:
                    summary.total_quantity = quantity  # Overwrite instead of add
                    summary.save()

                uploaded_count += 1

            except Warehouse.DoesNotExist:
                row["error"] = f"Warehouse '{warehouse_name}' not found"
                errors.append(row)
            except EducationalProduct.DoesNotExist:
                row["error"] = f"SKU '{sku_code}' not found"
                errors.append(row)
            except Exception as e:
                row["error"] = str(e)
                errors.append(row)

        # Handle errors
        if errors:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Errors"

            # Write headers
            headers = list(errors[0].keys())
            ws.append(headers)
            for row in errors:
                ws.append([row.get(h, '') for h in headers])

            for col in ws.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 2

            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename=upload_errors.xlsx'
            wb.save(response)
            return response

        messages.success(request, f"{uploaded_count} order summary rows uploaded successfully.")
        return redirect("upload_order_summary")

    return render(request, "orders/upload_order_summary.html")



# ✅ Dashboard View for Inventory Planning


def order_inventory_dashboard(request):
    selected_warehouse = request.GET.get('warehouse')
    selected_grade = request.GET.get('grade')
    selected_sku = request.GET.get('sku')

    # Thresholds
    thresholds = ProcurementThreshold.objects.first()
    low = thresholds.low_threshold if thresholds else 75
    medium = thresholds.medium_threshold if thresholds else 85

    # === ORDER DATA ===
    order_qs = OrderSummary.objects.select_related('warehouse', 'sku')

    if selected_warehouse:
        order_qs = order_qs.filter(warehouse_id=selected_warehouse)
    if selected_grade and selected_grade.lower() != "none":
        order_qs = order_qs.filter(grade__iexact=selected_grade.strip())
    if selected_sku:
        order_qs = order_qs.filter(sku__sku__icontains=selected_sku)

    order_map = {
        (o.warehouse.id, o.sku.sku): o
        for o in order_qs
        if o.warehouse and o.sku
    }

    # === INVENTORY DATA ===
    inventory_qs = InventoryItem.objects.select_related('product', 'warehouse')
    if selected_warehouse:
        inventory_qs = inventory_qs.filter(warehouse_id=selected_warehouse)
    if selected_sku:
        inventory_qs = inventory_qs.filter(product__sku__icontains=selected_sku)

    inventory_map = {
        (i.warehouse.id, i.product.sku): i
        for i in inventory_qs
        if i.warehouse and i.product
    }

    dashboard_data = []

    for key in order_map.keys():
        order = order_map[key]
        inventory = inventory_map.get(key)

        warehouse_name = order.warehouse.name
        grade = order.grade
        product_name = order.sku.product_description
        sku = order.sku.sku

        orders_received = order.total_quantity or 0
        stock_available = inventory.quantity_in_stock if inventory else 0

        percent_fulfilled = (orders_received / stock_available * 100) if stock_available > 0 else 0
        shortage = max(0, orders_received - stock_available)

        if orders_received == 0 and stock_available > 0:
            status = "Stock Available, No Orders"
        elif orders_received > 0 and stock_available == 0:
            status = "Need to Procure"
        elif percent_fulfilled < low:
            status = "OK"
        elif low <= percent_fulfilled < medium:
            status = "Yet to Procure"
        else:
            status = "Mandatory to Procure"

        dashboard_data.append({
            'warehouse': warehouse_name,
            'grade': grade,
            'sku': sku,
            'product': product_name,
            'orders_received': orders_received,
            'stock_available': stock_available,
            'shortage': shortage,
            'percent_fulfilled': round(percent_fulfilled, 1),
            'status': status,
            'alert': status != "OK"
        })

    # === FILTER VALUES ===
    warehouses = Warehouse.objects.all()
    grades = OrderSummary.objects.values_list('grade', flat=True).distinct().order_by('grade')
    if not grades.exists():
        grades = StudentOrder.objects.values_list('student_class', flat=True).distinct().order_by('student_class')

    return render(request, 'orders/dashboard.html', {
        'summary': dashboard_data,
        'warehouses': warehouses,
        'grades': grades,
        'selected_warehouse': selected_warehouse,
        'selected_grade': selected_grade,
        'selected_sku': selected_sku,
        'page_title': 'Inventory Planning Dashboard',
        'thresholds': {
            'low': low,
            'medium': medium
        }
    })  



@login_required
def update_thresholds(request):
    threshold = ProcurementThreshold.objects.filter(is_active=True).first()
    if request.method == 'POST':
        ok_threshold = int(request.POST.get('ok_threshold'))
        mandatory_threshold = int(request.POST.get('mandatory_threshold'))

        if threshold:
            threshold.low_threshold = ok_threshold
            threshold.medium_threshold = mandatory_threshold
            threshold.save()
        else:
            ProcurementThreshold.objects.create(
                low_threshold=ok_threshold,
                medium_threshold=mandatory_threshold,
                is_active=True
            )
        messages.success(request, "Thresholds updated successfully.")
        return redirect('order_inventory_dashboard')

    return render(request, 'orders/update_thresholds.html', {
        'threshold': threshold
    })


def order_summary_view(request):
    selected_warehouses = [w for w in request.GET.getlist('warehouse') if w]
    selected_grades = [g for g in request.GET.getlist('grade') if g]
    selected_volumes = [v for v in request.GET.getlist('volume') if v]
    selected_categories = [c for c in request.GET.getlist('category') if c]
    selected_subcategories = [s for s in request.GET.getlist('subcategory') if s]

    order_qs = OrderSummary.objects.select_related('sku', 'warehouse')

    if selected_warehouses:
        order_qs = order_qs.filter(warehouse_id__in=selected_warehouses)
    if selected_grades:
        order_qs = order_qs.filter(grade__in=selected_grades)
    if selected_volumes:
        order_qs = order_qs.filter(sku__volume__in=selected_volumes)
    if selected_categories:
        order_qs = order_qs.filter(sku__category__in=selected_categories)
    if selected_subcategories:
        order_qs = order_qs.filter(sku__sub_category__in=selected_subcategories)

    print(order_qs.query)

    # # Build a map of (warehouse_id, product_id) → stock quantity
    # inventory_qs = InventoryItem.objects.values('warehouse_id', 'product_id', 'quantity_in_stock')
    # inventory_map = {
    #     (item['warehouse_id'], item['product_id']): item['quantity_in_stock']
    #     for item in inventory_qs
    # }

    # Build a filtered inventory map only for the relevant warehouse + SKU IDs
    relevant_keys = set((order.warehouse.id, order.sku.id) for order in order_qs)

    if relevant_keys:
        inventory_qs = InventoryItem.objects.filter(
            reduce(lambda q1, q2: q1 | q2, [
                Q(warehouse_id=w_id, product_id=p_id)
                for (w_id, p_id) in relevant_keys
            ])
        ).values('warehouse_id', 'product_id', 'quantity_in_stock')
    else:
        inventory_qs = InventoryItem.objects.none()

    inventory_map = {
        (item['warehouse_id'], item['product_id']): item['quantity_in_stock']
        for item in inventory_qs
    }
 

    LOW_THRESHOLD = 75
    MEDIUM_THRESHOLD = 85

    # Add stock info to each OrderSummary object
    for order in order_qs:
        #print(f"LOOKUP: ({order.warehouse.id}, {order.sku.id})")
        #print(f"{order.warehouse} | {order.grade} | {order.sku.sku} | Qty: {order.total_quantity} | Stock: {getattr(order, 'stock_available', 0)}")
        key = (order.warehouse.id, order.sku.id)
        stock_available = inventory_map.get(key, 0)
        order.stock_available = stock_available

        orders_received = order.total_quantity

        percent_fulfilled = (orders_received / stock_available * 100) if stock_available > 0 else 0
        shortage = max(0, orders_received - stock_available)

        if orders_received == 0 and stock_available > 0:
            status = "Stock Available, No Orders"
        elif orders_received > 0 and stock_available == 0:
            status = "Need to Procure"
        elif percent_fulfilled < LOW_THRESHOLD:
            status = "OK"
        elif LOW_THRESHOLD <= percent_fulfilled < MEDIUM_THRESHOLD:
            status = "Yet to Procure"
        else:
            status = "Mandatory to Procure"

        # Attach to object (so you can use in template)
        order.percent_fulfilled = round(percent_fulfilled, 2)
        order.shortage = shortage
        order.procurement_status = status


    if request.GET.get('download') == '1':
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = 'attachment; filename="order_summary.csv"'

            writer = csv.writer(response)
            writer.writerow([
                'Warehouse', 'SKU', 'Description', 'Subcategory', 'Grade',
                'Volume', 'Order Quantity', 'Stock Available', '% Fulfilled',
                'Shortage', 'Procurement Status'
            ])
            for order in order_qs:
                writer.writerow([
                    order.warehouse.name,
                    order.sku.sku,
                    order.sku.product_description,
                    order.sku.sub_category,
                    order.grade,
                    order.sku.volume,
                    order.total_quantity,
                    order.stock_available,
                    order.percent_fulfilled,
                    order.shortage,
                    order.procurement_status,
                ])

            return response

    # Filter options
    warehouses = Warehouse.objects.all()
    grades = OrderSummary.objects.all().values_list('grade', flat=True).distinct()
    volumes = EducationalProduct.objects.all().values_list('volume', flat=True).distinct()
    categories = EducationalProduct.objects.all().values_list('category', flat=True).distinct()
    subcategories = EducationalProduct.objects.all().values_list('sub_category', flat=True).distinct()

    print("Filters applied:")
    print("Warehouses:", selected_warehouses)
    print("Grades:", selected_grades)
    print("Volumes:", selected_volumes)
    print("Categories:", selected_categories)
    print("Subcategories:", selected_subcategories)

    context = {
        'order_summaries': order_qs,
        'warehouses': warehouses,
        'grades': grades,
        'volumes': volumes,
        'categories': categories,
        'subcategories': subcategories,
        'selected_warehouses': selected_warehouses,
        'selected_grades': selected_grades,
        'selected_volumes': selected_volumes,
        'selected_categories': selected_categories,
        'selected_subcategories': selected_subcategories,
        'page_title': 'Order Summary View',
    }

    return render(request, 'orders/order_summary_filtered.html', context)

# Utility function for export
def export_order_summary(queryset, export_type):
    if export_type == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="order_summary.csv"'
        writer = csv.writer(response)
        writer.writerow(['SKU', 'Description', 'Segment', 'Year', 'Category', 'Subcategory', 'Grade', 'Volume', 'Unit', 'Publisher'])

        for obj in queryset:
            product = obj.sku
            writer.writerow([
                product.sku,
                product.product_description,
                product.segment,
                product.year,
                product.category,
                product.sub_category,
                obj.grade,
                product.volume,
                product.unit,
                product.publisher
            ])
        return response

    elif export_type == 'excel':
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="order_summary.xlsx"'
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['SKU', 'Description', 'Segment', 'Year', 'Category', 'Subcategory', 'Grade', 'Volume', 'Unit', 'Publisher'])

        for obj in queryset:
            product = obj.sku
            ws.append([
                product.sku,
                product.product_description,
                product.segment,
                product.year,
                product.category,
                product.sub_category,
                obj.grade,
                product.volume,
                product.unit,
                product.publisher
            ])

        wb.save(response)
        return response