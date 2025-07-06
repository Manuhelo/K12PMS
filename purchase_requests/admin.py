from django.contrib import admin
from .models import PurchaseRequest, RequestItem, PurchaseRequestStatusHistory
from products.models import EducationalProduct
from .forms import PurchaseRequestUploadForm, RequestItemCSVUploadForm
import openpyxl
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
import io
from django.contrib.auth import get_user_model
from django.urls import path, reverse
from django.shortcuts import render, redirect
from django.contrib import messages
from django.utils import timezone
from django.utils.html import format_html
from django.template.loader import render_to_string
from django.http import HttpResponseRedirect
from django.shortcuts import render, get_object_or_404
from django import forms    
from django.db.models import Sum
from django.utils.safestring import mark_safe
import csv
from io import TextIOWrapper
from procurement.models import VendorQuotation, PurchaseOrder, DeliveryTracking, RFQItem, RFQ, VendorBid  # assumed model locations
from django.template.response import TemplateResponse



User = get_user_model()


class RequestItemInline(admin.TabularInline):
    model = RequestItem
    extra = 0  # No extra empty rows
    raw_id_fields = ['product']  # Use input + search widget instead of dropdown
    readonly_fields = ['get_category', 'get_subcategory', 'get_grade']
    fields = ['product', 'quantity', 'get_category', 'get_subcategory', 'get_grade']

    def get_category(self, obj):
        return obj.product.category
    get_category.short_description = 'Category'

    def get_subcategory(self, obj):
        return obj.product.sub_category
    get_subcategory.short_description = 'Subcategory'

    def get_grade(self, obj):
        return obj.product.grade
    get_grade.short_description = 'Grade'

    def has_add_permission(self, request, obj=None):
        if not obj:  # While creating a new PurchaseRequest
            return True
        return (
            request.user.role == 'DepartmentUser' and
            obj.status in ['Draft', 'Rework Required']
        )

    def has_change_permission(self, request, obj=None):
        if not obj:
            return True
        return (
            request.user.role == 'DepartmentUser' and
            obj.status in ['Draft', 'Rework Required']
        )

    def has_delete_permission(self, request, obj=None):
        if not obj:
            return True
        return (
            request.user.role == 'DepartmentUser' and
            obj.status in ['Draft', 'Rework Required']
        )


@admin.register(PurchaseRequest)
class PurchaseRequestAdmin(admin.ModelAdmin):
    list_display = ('request_number', 'requested_by','segment', 'description', 'created_at', 'status', 'item_count', 'total_quantity','status_history_preview')
    change_list_template = "admin/purchase_requests/purchase_request_changelist.html"
    actions = ['submit_request','mark_as_reviewed', 'send_back_for_rework','export_with_items_to_excel']
    inlines = [RequestItemInline]


    def item_count(self, obj):
        return obj.request_items.count()
    item_count.short_description = 'Items Count'

    def total_quantity(self, obj):
        return obj.request_items.aggregate(total=Sum('quantity'))['total'] or 0
    total_quantity.short_description = 'Total Quantity'


    def status_history_preview(self, obj):
        updates = obj.status_history.all()
        if not updates.exists():
            return "No updates"
        
        popup_id = f"popup-{obj.id}"
        
        # Create a clickable link
        html = f'''
                <a href="#" onclick="togglePopup('{popup_id}'); return false;">
                    Status History
                </a>
                <div id="{popup_id}" style="display:none; background:#f9f9f9;color: #000;  padding:10px; margin-top:5px; border:1px solid #ccc;border-radius: 5px;
        max-width: 400px;
        box-shadow: 0 2px 8px rgba(0,0,0,0.2);">
                    <ul style="margin:0; padding:0; list-style-type:none;">
                        {''.join(f"<li><b>{u.changed_by}</b> ➝ <i>{u.new_status}</i> - <span style='color: blue;'>{u.changed_at.strftime('%Y-%m-%d %H:%M')}</span></li>" for u in updates)}
                    </ul>
                </div>
            '''
        return format_html(html)

    status_history_preview.short_description = 'Status History'

    class Media:
        js = ('admin/js/status_popup.js',)
    
    def formfield_for_choice_field(self, db_field, request, **kwargs):
        if db_field.name == 'status':
            user = request.user
            original_choices = db_field.choices

            if user.role == 'ProcurementManager':
                allowed_choices = ['Under Review', 'Rework Required', 'Reviewed']
                kwargs['choices'] = [choice for choice in original_choices if choice[0] in allowed_choices]

            elif user.role == 'DepartmentUser':
                allowed_choices = ['Draft', 'Submitted']
                kwargs['choices'] = [choice for choice in original_choices if choice[0] in allowed_choices]

            else:
                # For all other users, remove status field
                kwargs['choices'] = []

        return super().formfield_for_choice_field(db_field, request, **kwargs)
    
    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        if db_field.name == "requested_by":
            queryset = kwargs.get("queryset", User.objects.all())
            kwargs["queryset"] = queryset.filter(id=request.user.id)
        return super().formfield_for_foreignkey(db_field, request, **kwargs)


    def change_view(self, request, object_id, form_url='', extra_context=None):
        extra_context = extra_context or {}
        export_url = reverse('admin:purchase_request_export', args=[object_id])
        extra_context['export_url'] = export_url
        return super().change_view(request, object_id, form_url, extra_context=extra_context)
    

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('upload/', self.admin_site.admin_view(self.upload_csv), name='purchase_request_upload'),
            path('export/<int:pk>/', self.admin_site.admin_view(self.export_excel), name='purchase_request_export'),
            path('dashboard/', self.admin_site.admin_view(self.dashboard_view), name='purchase_request_dashboard'),
        ]
        return custom_urls + urls
    
    def get_queryset(self, request):
        qs = super().get_queryset(request)

        print("request.user:", request.user, type(request.user))

        if request.user.role == 'DepartmentUser':
            return qs.filter(requested_by=request.user)
        elif request.user.role == 'ProcurementManager':
            # Exclude Drafts
            return qs.exclude(status='Draft')
            # return qs.filter(status__in=['Submitted', 'Under Review'])
        return qs

    def get_readonly_fields(self, request, obj=None):
        if not obj:
            return []

        user = request.user
        if user.role == 'DepartmentUser':
            if obj.status in ['Submitted', 'Under Review', 'Reviewed']:
                return [f.name for f in self.model._meta.fields if f.name != 'status']
        elif user.role == 'ProcurementManager':
            # Let PM edit only the status
            return [f.name for f in self.model._meta.fields if f.name not in ['status', 'remarks']]
        return []

    def has_change_permission(self, request, obj=None):
        user = request.user
        if user.role == 'DepartmentUser' and obj:
            return obj.status in ['Draft', 'Rework Required']
        
        # ProcurementManager should not be able to change anything
        if user.role == 'ProcurementManager':
            return True
        return super().has_change_permission(request, obj)
    
    # ✅ Add this method
    def submit_request(self, request, queryset):
        updated = queryset.update(status='Submitted')
        self.message_user(request, f"{updated} request(s) submitted for review.")
    submit_request.short_description = "Submit to Procurement for Review"

    def mark_as_reviewed(self, request, queryset):
        updated = queryset.update(status='Reviewed')
        self.message_user(request, f"{updated} request(s) marked as Reviewed.")
    mark_as_reviewed.short_description = "Approve Request (Mark as Reviewed)"

    def send_back_for_rework(self, request, queryset):
        updated = queryset.update(status='Rework Required')
        self.message_user(request, f"{updated} request(s) sent back for rework.")
    send_back_for_rework.short_description = "Send Back to Department for Rework"

    def upload_csv(self, request):
        if request.method == "POST":
            form = PurchaseRequestUploadForm(request.POST, request.FILES)
            if form.is_valid():
                try:
                    file = request.FILES['file']
                    filename = file.name.lower()

                    rows = []

                    if filename.endswith('.csv'):
                        decoded_file = TextIOWrapper(file, encoding='utf-8')
                        reader = csv.DictReader(decoded_file)
                        for row in reader:
                            rows.append({
                                'sku': row['sku'].strip().upper(),
                                'quantity': row['quantity'],
                                'remarks': row.get('remarks', '')
                            })
                    elif filename.endswith(('.xls', '.xlsx')):
                        wb = openpyxl.load_workbook(file)
                        sheet = wb.active
                        headers = [cell.value.strip().lower() for cell in sheet[1]]

                        for row in sheet.iter_rows(min_row=2, values_only=True):
                            row_dict = dict(zip(headers, row))
                            rows.append({
                                'sku': str(row_dict['sku']).strip().upper(),
                                'quantity': row_dict.get('quantity'),
                                'remarks': row_dict.get('remarks', '')
                            })
                    else:
                        messages.error(request, "Unsupported file type. Please upload a CSV or Excel file.")
                        return redirect("..")
                except Exception as e:
                    messages.error(request, f"File processing error: {str(e)}")
                    return redirect("..")
                
                # Validate SKUs
                skus = [r['sku'] for r in rows]

                # Normalize DB SKUs as well
                existing_products_qs = EducationalProduct.objects.filter(sku__in=skus)
                existing_products = {p.sku.strip().upper(): p for p in existing_products_qs}

                missing_skus = [sku for sku in skus if sku not in existing_products]

                if missing_skus:
                    messages.error(request, f"Purchase request not created. Missing SKUs: {', '.join(missing_skus)}")
                    return redirect("..")

                # Create purchase request
                now = timezone.now()
                user = request.user
                segment = form.cleaned_data['segment']  # from form input
                description = form.cleaned_data['description']
                remarks = form.cleaned_data['remarks']
                request_number = f"REQ-{now.strftime('%Y%m%d')}-{PurchaseRequest.objects.count() + 1:03d}"

                purchase_request = PurchaseRequest.objects.create(
                    requested_by=user,
                    segment=segment,  # Add this line
                    description=description,
                    status='Draft',
                    remarks=remarks,
                    request_number=request_number
                )

                for row in rows:
                    sku = row['sku']

                    try:
                        quantity = int(str(row['quantity']).replace(',', '').strip())
                    except ValueError:
                        messages.error(request, f"Invalid quantity value: {row.get('quantity')} in row: {row.to_dict()}")
                        return redirect("..")

                    product = existing_products.get(sku)  # Now it's safe!
                    RequestItem.objects.create(
                        purchase_request=purchase_request,
                        product=product,
                        quantity=quantity,
                        remarks=row.get('remarks', '')
                    )

                messages.success(request, f"Purchase Request {purchase_request.request_number} created successfully.")
                return redirect("..")

        else:
            form = PurchaseRequestUploadForm()

        sample_data = [
        {"title": "Grade 1_English Workbook_Vol 2 (25)", "sku": "4000008632", "quantity": 9450},
        {"title": "Grade 1_Maths Textbook_Vol 2 (25)", "sku": "4000008821", "quantity": 9450},
        {"title": "Grade 1_Maths Workbook_Vol 2 (25)", "sku": "4000008785", "quantity": 9450},
        {"title": "Grade 1_Trilingual Hin-Kan_Rani's Wish_Vol 2 (25)", "sku": "4000008758", "quantity": 2973},
        {"title": "Grade 1_Trilingual Hin-Mar_Rani's Wish_Vol 2 (25)", "sku": "4000008766", "quantity": 3447},
        ]
        return render(request, "admin/purchase_requests/purchase_request_upload.html", {"form": form, "sample_data": sample_data})
    
    def export_excel(self, request, pk):
        from .models import PurchaseRequest  # adjust if needed

        pr = PurchaseRequest.objects.get(pk=pk)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Purchase Request"

        headers = [
            'Request ID', 'Status','Segment', 'Requested By', 'Created At',
            'Item ID', 'Product Name', 'Category', 'Subcategory', 'Grade', 'Quantity'
        ]
        ws.append(headers)

        for item in pr.request_items.all():
            product = item.product
            ws.append([
                pr.request_number,
                pr.status,
                pr.segment,  # Add this
                pr.requested_by.username,
                pr.created_at.strftime("%Y-%m-%d %H:%M"),
                product.sku,
                product.product_description,
                product.category,
                product.sub_category,
                product.grade,
                item.quantity
            ])

        for col in ws.columns:
            max_length = max(len(str(cell.value) if cell.value else "") for cell in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=purchase_request_{pk}.xlsx'
        wb.save(response)
        return response
    

    def dashboard_view(self, request):
        data = []

        all_requests = PurchaseRequest.objects.all().order_by('-created_at')

        for pr in all_requests:
            # Count items
            item_count = pr.request_items.count()
            # Get related RFQs for the purchase request first
            rfqs = pr.rfqs.all()

            # Get the correct count based on unique (rfq, vendor, submission_group)
            quotation_count = VendorBid.objects.filter(rfq__in=rfqs).values(
                'rfq', 'vendor', 'submission_group'
            ).distinct().count()

            approved_bids = VendorBid.objects.filter(rfq__purchase_request=pr, status="Approved")
            approved_vendors = approved_bids.values_list('vendor__name', flat=True).distinct()


            pos = PurchaseOrder.objects.filter(vendor_bid__rfq__in=rfqs)
            po_numbers = pos.values_list('po_number', flat=True)
            delivery_dates = DeliveryTracking.objects.filter(purchase_order__in=pos).values_list('estimated_delivery', flat=True)


            data.append({
                'request_number': pr.request_number,
                'item_count': item_count,
                'quotation_count': quotation_count,
                'approved_vendors': ", ".join(approved_vendors),
                'po_numbers': ", ".join(po_numbers),
                'delivery_dates': ", ".join([d.strftime('%Y-%m-%d') for d in delivery_dates]),
            })

        context = dict(
            self.admin_site.each_context(request),
            title="Purchase Request Dashboard",
            total_requests=PurchaseRequest.objects.count(),
            submitted_requests=PurchaseRequest.objects.filter(status='Submitted').count(),
            reviewed_requests=PurchaseRequest.objects.filter(status='Reviewed').count(),
            dashboard_data=data,
        )
        return TemplateResponse(request, "admin/purchase_requests/dashboard.html", context)

    def save_model(self, request, obj, form, change):
        obj._changed_by = request.user
        super().save_model(request, obj, form, change)

# admin.site.register(PurchaseRequest)
# admin.site.register(RequestItem)
@admin.register(PurchaseRequestStatusHistory)
class PurchaseRequestStatusHistoryAdmin(admin.ModelAdmin):
    list_display = ('purchase_request', 'old_status', 'new_status','changed_by', 'changed_at')




@admin.register(RequestItem)
class RequestItemAdmin(admin.ModelAdmin):
    list_display = ('purchase_request', 'product', 'quantity','category', 
        'subcategory', 
        'grade')
    
    def get_queryset(self, request):
        qs = super().get_queryset(request)
        print("request.user:", request.user, type(request.user))
        if request.user.role == 'ProcurementManager':
            # Show items only for non-Draft requests
            return qs.exclude(purchase_request__status='Draft')
        elif request.user.role == 'DepartmentUser':
            return qs.filter(purchase_request__requested_by=request.user)
        return qs
    
    def category(self, obj):
        return obj.product.category if obj.product else None

    def subcategory(self, obj):
        return obj.product.sub_category if obj.product else None

    def grade(self, obj):
        return obj.product.grade if obj.product else None

    def has_change_permission(self, request, obj=None):
        if request.user.role == 'ProcurementManager':
            return False
        
        if obj and obj.purchase_request.status not in ['Draft', 'Rework Required']:
            return False
        
        return super().has_change_permission(request, obj)

    def has_add_permission(self, request):

        if request.user.role == 'ProcurementManager':
            return False
        
        # Only allow adding if purchase request ID is in the URL and it's in Draft
        purchase_request_id = request.GET.get('purchase_request__id__exact')
        if purchase_request_id:
            try:
                pr = PurchaseRequest.objects.get(id=purchase_request_id)
                return pr.status in ['Draft', 'Rework Required']
            except PurchaseRequest.DoesNotExist:
                return False
        return True

    def has_delete_permission(self, request, obj=None):
        if request.user.role == 'ProcurementManager':
            return False
        
        if obj and obj.purchase_request.status not in ['Draft', 'Rework Required']:
            return False
        return super().has_delete_permission(request, obj)
    

#admin.site.register(PurchaseOrder)
admin.site.register(DeliveryTracking)
    
