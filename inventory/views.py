from django.shortcuts import render,redirect, get_object_or_404
from procurement.models import PurchaseOrder, VendorQuotation
from .models import *
from django.contrib import messages
from django.urls import reverse
from .forms import *
from products.models import *
import csv
from django.http import HttpResponse
from django.core.files.uploadedfile import InMemoryUploadedFile
import io
from django.db.models import Sum, Q
from xhtml2pdf import pisa
from io import BytesIO
from django.template.loader import get_template
from .utils.decorators import warehouse_roles_required 
from django.contrib.auth.decorators import login_required, user_passes_test
from users.models import CustomUser  # adjust path if needed
from django.utils.timezone import now
import openpyxl
from django.utils import timezone
from openpyxl import Workbook


def scan_po(request):
    if request.method == 'POST':
        po_number = request.POST.get('po_number')
        try:
            # If PO exists, redirect to details view
            PurchaseOrder.objects.get(po_number=po_number)
            return redirect('receive_goods', po_number=po_number)
        except PurchaseOrder.DoesNotExist:
            return render(request, 'inventory/scan_po.html', {
                'error': 'Purchase Order not found.'
            })

    return render(request, 'inventory/scan_po.html')



def po_details_by_number(request, po_number):
    po = get_object_or_404(PurchaseOrder, po_number=po_number)
    
    vendor_bid = po.vendor_bid
    quotations = VendorQuotation.objects.filter(vendor_bid=vendor_bid).select_related(
        'rfq_item__request_item__product'
    )

    item_details = []
    for quote in quotations:
        req_item = quote.rfq_item.request_item
        item_details.append({
            'product': req_item.product.product_description if req_item.product else 'Unknown',
            'quantity': req_item.quantity,
            'quoted_price': quote.quoted_price,
            'lead_time_days': quote.lead_time_days
        })

    return render(request, 'admin/procurement/po_details.html', {
        'po': po,
        'vendor': vendor_bid.vendor if vendor_bid else None,
        'item_details': item_details,
    })



def receive_po_view(request, po_number):
    po = get_object_or_404(PurchaseOrder, po_number=po_number)
    vendor_bid = po.vendor_bid

    quotations = VendorQuotation.objects.filter(vendor_bid=vendor_bid).select_related(
        'rfq_item__request_item__product'
    )

    item_details = []
    for quote in quotations:
        req_item = quote.rfq_item.request_item
        item_details.append({
            'product_id': req_item.product.id,
            'product_name': req_item.product.product_description,
            'quantity': req_item.quantity,
            'quoted_price': quote.quoted_price,
            'lead_time_days': quote.lead_time_days
        })

    warehouses = Warehouse.objects.all()

    if request.method == 'POST':
        warehouse_id = request.POST.get('warehouse')
        if not warehouse_id:
            messages.error(request, "Please select a warehouse.")
            return redirect(request.path)

        warehouse = get_object_or_404(Warehouse, id=warehouse_id)

        # Create GoodsReceipt
        grn = GoodsReceipt.objects.create(
            purchase_order=po,
            warehouse=warehouse,
            received_by=request.user,
            remarks='Auto GRN generated'
        )

        # Loop through items and create GoodsReceiptItem for each with quantity input
        for item in item_details:
            qty_key = f"qty_{item['product_id']}"
            qty_received = int(request.POST.get(qty_key, 0))

            if qty_received > 0:
                GoodsReceiptItem.objects.create(
                    receipt=grn,
                    product_id=item['product_id'],
                    quantity_received=qty_received
                )

        messages.success(request, f"Goods Receipt created for PO {po.po_number}.")
        return redirect('admin:index')  # or any page you want to redirect to

    return render(request, 'inventory/receive_goods.html', {
        'po': po,
        'vendor': vendor_bid.vendor if vendor_bid else None,
        'item_details': item_details,
        'warehouses': warehouses,
    })


def not_authorized(request):
    return render(request, 'inventory/403.html', {'page_title': 'Access Denied'})

@warehouse_roles_required
def add_warehouse(request):
    form = WarehouseForm(request.POST or None)
    context = {
        'form': form,
        'page_title': 'Add Warehouse'
    }
    if request.method == 'POST':
        if form.is_valid():
            try:
                form.save()
                messages.success(request, "Successfully Added")
                return redirect(reverse('add_warehouse'))
            except Exception as e:
                messages.error(request, f"Could not add: {str(e)}")
        else:
            messages.error(request, "Invalid form data. Please check and try again.")
    return render(request, 'inventory/add_warehouse.html', context)

@warehouse_roles_required
def upload_warehouses(request):
    if request.method == 'POST' and request.FILES.get('file'):
        file = request.FILES['file']
        ext = file.name.split('.')[-1]

        try:
            warehouses = []

            if ext == 'csv':
                decoded_file = file.read().decode('utf-8').splitlines()
                reader = csv.DictReader(decoded_file)
                rows = list(reader)

            elif ext in ['xls', 'xlsx']:
                wb = openpyxl.load_workbook(file)
                ws = wb.active
                headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

                rows = []
                for row in ws.iter_rows(min_row=2, values_only=True):
                    row_dict = dict(zip(headers, row))
                    rows.append(row_dict)

            else:
                messages.error(request, "Unsupported file type.")
                return redirect('upload_warehouses')

            for row in rows:
                manager_username = row.get('manager')
                manager = None
                if manager_username:
                    manager = CustomUser.objects.filter(username=manager_username).first() or \
                              CustomUser.objects.filter(email=manager_username).first()

                warehouses.append(
                    Warehouse(
                        name=row['name'],
                        location=row.get('location', ''),
                        manager=manager
                    )
                )

            Warehouse.objects.bulk_create(warehouses)
            messages.success(request, f"{len(warehouses)} warehouses uploaded successfully.")

        except Exception as e:
            messages.error(request, f"Error processing file: {e}")

        return redirect('upload_warehouses')

    return render(request, 'inventory/upload_warehouses.html', {'page_title': 'Upload Warehouses'})

@warehouse_roles_required
def manage_warehouse(request):
    warehouses = Warehouse.objects.all()
    context = {
        'warehouses': warehouses,
        'page_title': 'Manage Warehouses'
    }
    return render(request, "inventory/manage_warehouse.html", context)

@warehouse_roles_required
def edit_warehouse(request, warehouse_id):
    warehouse = get_object_or_404(Warehouse, id=warehouse_id)
    form = WarehouseForm(request.POST or None, instance=warehouse)
    context = {
        'form': form,
        'warehouse_id': warehouse_id,
        'page_title': 'Edit Course'
    }
    if request.method == 'POST':
        if form.is_valid():
            form.save()
            messages.success(request, "Warehouse updated successfully.")
            return redirect('manage_warehouse')
        else:
            messages.error(request, "Could Not Update")

    return render(request, 'inventory/edit_warehouse.html', context)

@warehouse_roles_required
def delete_warehouse(request, warehouse_id):
    warehouse = get_object_or_404(Warehouse, id=warehouse_id)
    try:
        warehouse.delete()
        messages.success(request, "Warehouse deleted successfully!")
    except Exception:
        messages.error(
            request, "Sorry, this warehouse cant delete.")
    return redirect(reverse('manage_warehouse'))


def safe_int(value):
    try:
        return int(float(value)) if value not in [None, '', 'NaN', 'nan'] else 0
    except (ValueError, TypeError):
        return 0

def clean_sku(raw):
    try:
        return str(int(float(raw))).strip()  # Converts 4000008632.0 → '4000008632'
    except:
        return str(raw).strip()
    

def manage_grns(request):
    grns = GoodsReceipt.objects.all().order_by('-received_at')
    return render(request, 'inventory/grn_list.html', {'grns': grns, 'page_title': 'Manage GRNs'})


def grn_detail(request, grn_id):
    grn = get_object_or_404(GoodsReceipt, id=grn_id)
    items = grn.items.select_related('product')  # GoodsReceiptItem objects

    # Build a list of items with ordered quantity
    display_items = []

    total_ordered = 0
    total_received = 0
    total_damaged = 0
    total_pending = 0

    for item in items:
        try:
            po_item = POItem.objects.get(purchase_order=grn.purchase_order, product=item.product)
            ordered_qty = po_item.quantity_ordered
        except POItem.DoesNotExist:
            ordered_qty = 0

        received_qty = item.quantity_received or 0
        damaged_qty = item.damage_return or 0
        pending_qty = max(ordered_qty - received_qty, 0)

        total_ordered += ordered_qty or 0
        total_received += received_qty
        total_damaged += damaged_qty
        total_pending += pending_qty

        display_items.append({
            'sku': item.product.sku,
            'product': item.product.product_description,
            'ordered_qty': ordered_qty,
            'received_qty': item.quantity_received,
            'damage_qty': item.damage_return or 0,
            'pending_qty': pending_qty
        })

    return render(request, 'inventory/grn_detail.html', {
        'grn': grn,
        'items': display_items,
        'total_ordered': total_ordered,
        'total_received': total_received,
        'total_damaged': total_damaged,
        'total_pending': total_pending,
        'page_title': f'GRN Details for {grn.purchase_order.po_number}'
    })



def delete_grn(request, grn_id):
    grn = get_object_or_404(GoodsReceipt, id=grn_id)

    if request.method == "POST":
        grn.delete()
        messages.success(request, f"GRN for PO {grn.purchase_order.po_number} deleted successfully.")
        return redirect('grn_manage')  # URL name for GRN list page

    # Optional: Confirm page (or skip and use inline form)
    return render(request, 'inventory/grn_confirm_delete.html', {
        'grn': grn,
        'page_title': 'Confirm Delete GRN'
    })



def download_grn_detail(request, grn_id):
    grn = get_object_or_404(GoodsReceipt, id=grn_id)
    items = grn.items.select_related('product')

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="GRN_{grn_id}_detail.csv"'

    writer = csv.writer(response)
    writer.writerow([
        'GRN ID','SKU Code', 'Product Description', 'Ordered Qty',
        'Received Qty', 'Damaged Qty', 'Pending Qty'
    ])

    for item in items:
        try:
            po_item = POItem.objects.get(purchase_order=grn.purchase_order, product=item.product)
            ordered_qty = po_item.quantity_ordered
        except POItem.DoesNotExist:
            ordered_qty = 0

        received_qty = item.quantity_received
        damage_qty = item.damage_return or 0
        pending_qty = max(ordered_qty - received_qty, 0)

        writer.writerow([
            grn.id,
            item.product.sku,
            item.product.product_description,
            ordered_qty,
            received_qty,
            damage_qty,
            pending_qty
        ])

    return response


def download_grn_pdf(request, grn_id):
    grn = get_object_or_404(GoodsReceipt, id=grn_id)
    items = grn.items.select_related('product')

    display_items = []
    total_ordered = 0
    total_received = 0
    total_damaged = 0

    for item in items:
        try:
            po_item = POItem.objects.get(purchase_order=grn.purchase_order, product=item.product)
            ordered_qty = po_item.quantity_ordered
        except POItem.DoesNotExist:
            ordered_qty = 0

        received_qty = item.quantity_received
        damage_qty = item.damage_return or 0
        pending_qty = max(ordered_qty - received_qty, 0)

        display_items.append({
            'sku': item.product.sku,
            'product': item.product.product_description,
            'ordered_qty': ordered_qty,
            'received_qty': received_qty,
            'damage_qty': damage_qty,
            'pending_qty': pending_qty,
        })

        total_ordered += ordered_qty
        total_received += received_qty
        total_damaged += damage_qty

    total_pending = total_ordered - total_received

    context = {
        'grn': grn,
        'items': display_items,
        'total_ordered': total_ordered,
        'total_received': total_received,
        'total_damaged': total_damaged,
        'total_pending': total_pending,
    }

    template = get_template('inventory/grn_detail_pdf.html')
    html = template.render(context)

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename=GRN_{grn_id}_Report.pdf'

    pisa_status = pisa.CreatePDF(BytesIO(html.encode('utf-8')), dest=response)
    if pisa_status.err:
        return HttpResponse('Error generating PDF', status=500)
    return response



def inventory_list(request):
    # Get all filter options
    warehouses = Warehouse.objects.all()
    products = EducationalProduct.objects.all()
    # Get filters from GET request
    selected_warehouses = request.GET.getlist('warehouse')
    selected_grades = request.GET.getlist('grade')
    selected_categories = request.GET.getlist('category')
    selected_subcategories = request.GET.getlist('subcategory')
    selected_volumes = request.GET.getlist('volume')

    queryset = InventoryItem.objects.select_related('product')

    if selected_warehouses:
        queryset = queryset.filter(warehouse_id__in=selected_warehouses)

    if selected_grades:
        queryset = queryset.filter(product__grade__in=selected_grades)

    if selected_categories:
        queryset = queryset.filter(product__category__in=selected_categories)

    if selected_subcategories:
        queryset = queryset.filter(product__sub_category__in=selected_subcategories)

    if selected_volumes:
        queryset = queryset.filter(product__volume__in=selected_volumes)

    grouped_data = queryset.values(
        'product__sku', 'product__product_description','product__category','product__sub_category','product__grade','product__volume'
    ).annotate(total_qty=Sum('quantity_in_stock'))

    # Handle CSV Download
    if request.GET.get('download') == '1':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="inventory_summary.csv"'

        writer = csv.writer(response)
        writer.writerow(['SKU Code', 'Product Description','Category','Sub Category','Grade', 'Volume','Total Quantity'])

        for row in grouped_data:
            writer.writerow([row['product__sku'], row['product__product_description'],row['product__category'],row['product__sub_category'],row['product__grade'],row['product__volume'], row['total_qty']])
        return response

    context = {
        'grouped_data': grouped_data,
        'warehouses': Warehouse.objects.all(),
        'grades': EducationalProduct.objects.values_list('grade', flat=True).distinct(),
        'categories': EducationalProduct.objects.values_list('category', flat=True).distinct(),
        'subcategories': EducationalProduct.objects.values_list('sub_category', flat=True).distinct(),
        'volumes': EducationalProduct.objects.values_list('volume', flat=True).distinct(),
        'selected_warehouses': list(map(int, selected_warehouses)),
        'selected_grades': selected_grades,
        'selected_categories': selected_categories,
        'selected_subcategories': selected_subcategories,
        'selected_volumes': selected_volumes,
        'page_title': 'Inventory Summary',
    }

    return render(request, 'inventory/inventory_list.html', context)


def stock_history_list(request):
    history = StockHistory.objects.select_related('product', 'warehouse', 'changed_by').order_by('-changed_at')
    return render(request, 'inventory/stock_history.html', {
        'history': history,
        'page_title': 'Stock Change History'
    })


def upload_grn_items(request):
    if request.method == 'POST' and 'upload' in request.POST:
        po_id = request.POST.get('po_id')
        warehouse_id = request.POST.get('warehouse_id')
        warehouse = Warehouse.objects.get(id=warehouse_id)
        po = PurchaseOrder.objects.get(id=po_id)
        uploaded_file = request.FILES['file']

        preview_data = []
        errors = []

        # Read file using pandas
        try:
            if uploaded_file.name.endswith('.csv'):
                decoded_file = uploaded_file.read().decode('utf-8').splitlines()
                reader = csv.DictReader(decoded_file)
                rows = list(reader)
            elif uploaded_file.name.endswith(('.xls', '.xlsx')):
                wb = openpyxl.load_workbook(uploaded_file)
                ws = wb.active
                headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

                rows = []
                for row in ws.iter_rows(min_row=2, values_only=True):
                    rows.append(dict(zip(headers, row)))
            else:
                return render(request, 'inventory/grn_upload.html', {
                    'pos': PurchaseOrder.objects.all(),
                    'warehouses': Warehouse.objects.all(),
                    'error': 'Unsupported file type. Please upload .csv or .xlsx',
                    'page_title': 'Bulk Upload GRN'
                })
        except Exception as e:
            return render(request, 'inventory/grn_upload.html', {
                'pos': PurchaseOrder.objects.all(),
                'warehouses': Warehouse.objects.all(),
                'error': f'Error reading file: {str(e)}',
                'page_title': 'Bulk Upload GRN'
            })

        product_map = {
            p.sku: p for p in EducationalProduct.objects.all()
        }

        # Get product IDs from this PO's vendor_bid
        valid_product_ids = list(po.vendor_bid.quotations.values_list(
            'rfq_item__request_item__product_id', flat=True
        ))


        # Optional: Map product_id to PO quantity (assuming you store quantity somewhere)
        po_items_map = {   }
        quotations = po.vendor_bid.quotations.select_related('rfq_item__request_item')

        for q in quotations:
            product_id = q.rfq_item.request_item.product_id
            qty = q.rfq_item.request_item.quantity  # or use q.quantity if applicable
            po_items_map[product_id] = qty

        for row in rows:
            sku = clean_sku(row.get('SKU Code'))
            grn_qty = safe_int(row.get('GRN Quantity', 0))
            damage_qty = safe_int(row.get('Damage Quantity', 0))

            product = product_map.get(sku)

            if not product:
                errors.append(f"Invalid SKU: {sku}")
                continue

            if product.id not in valid_product_ids:
                errors.append(f"Product {product.product_description} is not part of the selected PO.")
                continue

            po_qty = po_items_map.get(product.id, 'N/A')

            preview_data.append({
                'product': product,
                'sku': sku,
                'grn_qty': grn_qty,
                'damage_qty': damage_qty,
                'po_qty': po_qty
            })

        return render(request, 'inventory/grn_preview.html', {
            'preview_data': preview_data,
            'errors': errors,
            'po': po,
            'warehouse': warehouse,
            'page_title': 'Preview GRN Items'
        })

    # GET request
    return render(request, 'inventory/grn_upload.html', {
        'pos': PurchaseOrder.objects.all(),
        'warehouses': Warehouse.objects.all(),
        'error': None,
        'page_title': 'Bulk Upload GRN'
    })


# confirm_grn_upload view
def confirm_grn_upload(request):
    if request.method == 'POST':
        po = get_object_or_404(PurchaseOrder, id=request.POST.get('po_id'))
        warehouse = get_object_or_404(Warehouse, id=request.POST.get('warehouse_id'))
        user = request.user

        sku_list = request.POST.getlist('sku[]')
        grn_qty_list = request.POST.getlist('grn_qty[]')
        damage_qty_list = request.POST.getlist('damage_qty[]')

        grn = GoodsReceipt.objects.create(
            purchase_order=po,
            warehouse=warehouse,
            received_by=user
        )

        quotations = po.vendor_bid.quotations.select_related('rfq_item__request_item__product')
        quotation_map = {
            str(q.rfq_item.request_item.product.sku): {
                'product': q.rfq_item.request_item.product,
                'ordered_qty': q.rfq_item.request_item.quantity
            }
            for q in quotations
        }

        for sku, qty, damage_qty in zip(sku_list, grn_qty_list, damage_qty_list):
            cleaned_sku = str(sku).strip().replace('.0', '')
            data = quotation_map.get(cleaned_sku)

            if not data:
                continue

            product = data['product']
            ordered_qty = data['ordered_qty']
            delta_received = int(qty)
            delta_damage = int(damage_qty)

            if delta_received > 0:
                po_item, _ = POItem.objects.get_or_create(
                    purchase_order=po,
                    product=product,
                    defaults={'quantity_ordered': ordered_qty}
                )

                GoodsReceiptItem.objects.create(
                    receipt=grn,
                    product=product,
                    quantity_received=delta_received,
                    damage_return=delta_damage
                )

                GRNRecord.objects.create(
                    po_item=po_item,
                    grn=grn,
                    quantity_received=delta_received,
                    damage_quantity=delta_damage
                )

        # ✅ Auto-mark PO as Delivered if all items received
        po_items = POItem.objects.filter(purchase_order=po)
        all_fulfilled = True
        for item in po_items:
            total_received = GoodsReceiptItem.objects.filter(
                receipt__purchase_order=po,
                receipt__warehouse=warehouse,
                product=item.product
            ).aggregate(total=Sum('quantity_received'))['total'] or 0

            if total_received < item.quantity_ordered:
                all_fulfilled = False

        if all_fulfilled:
            po.status = 'Delivered'
        else:
            po.status = 'Partially Received'
        po.save()

        return redirect('grn_manage')




def po_list_for_grn(request):
    pos = PurchaseOrder.objects.filter(status__in=['Approved', 'Partially Received', 'Delivered'])
    po_data = []

    for po in pos:
        po_items = POItem.objects.filter(purchase_order=po)
        total_ordered = po_items.aggregate(total=Sum('quantity_ordered'))['total'] or 0

        # ❗Filter GRNs by this PO and this PO's warehouse
        gr_items = GoodsReceiptItem.objects.filter(
            receipt__purchase_order=po,
            receipt__warehouse=po.Warehouse
        )

        capped_received = 0
        for item in po_items:
            received = gr_items.filter(product=item.product).aggregate(
                total=Sum('quantity_received'))['total'] or 0
            capped_received += min(received, item.quantity_ordered)

        progress = int((capped_received / total_ordered) * 100) if total_ordered else 0
        fully_received = capped_received >= total_ordered

        grns = GoodsReceipt.objects.filter(purchase_order=po, warehouse=po.Warehouse)
        grn_exists = grns.exists()
        first_grn = grns.first() if grn_exists else None

        po_data.append({
            'po': po,
            'progress': progress,
            'fully_received': fully_received,
            'grn_exists': grn_exists,
            'grn_id': first_grn.id if first_grn else None
        })

    return render(request, 'inventory/po_list_for_grn.html', {'po_data': po_data})

def po_grn_entry(request, po_id):
    po = get_object_or_404(PurchaseOrder, id=po_id)
    warehouse = po.Warehouse
    quotations = po.vendor_bid.quotations.select_related('rfq_item__request_item__product')

    preview_data = []
    for q in quotations:
        product = q.rfq_item.request_item.product
        po_qty = q.rfq_item.request_item.quantity

        # Calculate already received quantity for this product in this PO and warehouse
        already_received = GoodsReceiptItem.objects.filter(
            receipt__purchase_order=po,
            receipt__warehouse=warehouse,
            product=product
        ).aggregate(total=Sum('quantity_received'))['total'] or 0

        preview_data.append({
            'sku': product.sku,
            'product': product,
            'po_qty': po_qty,
            'already_received': already_received,
        })

    return render(request, 'inventory/po_grn_preview.html', {
        'po': po,
        'warehouse': warehouse,
        'preview_data': preview_data,
        'errors': [],
        'page_title': f'GRN Entry for PO {po.po_number}'
    })


def po_grn_detail(request, po_id):
    po = get_object_or_404(PurchaseOrder, id=po_id)
    warehouse = po.Warehouse

    po_items = POItem.objects.filter(purchase_order=po)
    preview_data = []

    total_ordered = 0
    total_received = 0
    total_pending = 0

    for item in po_items:
        product = item.product
        ordered_qty = item.quantity_ordered

        agg = GoodsReceiptItem.objects.filter(
            receipt__purchase_order=po,
            receipt__warehouse=warehouse,
            product=product
        ).aggregate(
            received_qty=Sum('quantity_received'),
        )

        received_qty = agg['received_qty'] or 0
        pending_qty = max(0, ordered_qty - received_qty)

        total_ordered += ordered_qty
        total_received += received_qty
        total_pending += pending_qty

        preview_data.append({
            'sku': product.sku,
            'product': product,
            'ordered_qty': ordered_qty,
            'received_qty': received_qty,
            'pending_qty': pending_qty
        })

    latest_grn = po.goodsreceipt_set.order_by('-received_at').first()

    return render(request, 'inventory/po_grn_detail.html', {
        'po': po,
        'warehouse': warehouse,
        'received_by': latest_grn.received_by if latest_grn else None,
        'received_at': latest_grn.received_at if latest_grn else None,
        'preview_data': preview_data,
        'total_ordered': total_ordered,
        'total_received': total_received,
        'total_pending': total_pending,
        'page_title': f'GRN Summary for PO {po.po_number}'
    })

def po_fulfillment_report(request, po_id):
    po = PurchaseOrder.objects.get(id=po_id)
    po_items = POItem.objects.filter(purchase_order=po)

    report = []
    for item in po_items:
        received = item.grn_records.aggregate(total=Sum('quantity_received'))['total'] or 0
        pending = item.quantity_ordered - received
        report.append({
            'product': item.product.product_description,
            'ordered': item.quantity_ordered,
            'received': received,
            'pending': pending
        })

    return render(request, "inventory/po_fulfillment.html", {
        'report': report,
        'po': po
    })


# @login_required
# @warehouse_roles_required
def create_stock_request(request):
    warehouses = Warehouse.objects.all()
    products = EducationalProduct.objects.all()

    if request.method == 'POST':
        warehouse_id = request.POST.get('warehouse')
        remarks = request.POST.get('remarks')

        product_ids = request.POST.getlist('product')
        quantities = request.POST.getlist('quantity')

        print("DEBUG -- product_ids:", product_ids)
        print("DEBUG -- quantities:", quantities)

        stock_request = StockRequest.objects.create(
            requesting_warehouse_id=warehouse_id,
            requested_by=request.user,
            remarks=remarks
        )

        for prod_id, qty in zip(product_ids, quantities):
            try:
                qty_int = int(qty)
                if qty_int > 0:
                    StockRequestItem.objects.create(
                        stock_request=stock_request,
                        product_id=prod_id,
                        quantity_requested=qty_int
                    )
            except ValueError:
                continue  # skip if invalid

        return redirect('stock_request_list')

    return render(request, 'inventory/create_stock_request.html', {
        'warehouses': warehouses,
        'products': products
    })

def view_stock_request(request, pk):
    stock_request = get_object_or_404(StockRequest, pk=pk)
    return render(request, 'inventory/view_stock_request.html', {
        'stock_request': stock_request,
        'items': stock_request.items.all(),
        'page_title': 'Stock Request Detail'
    })

def delete_stock_request(request, pk):
    stock_request = get_object_or_404(StockRequest, pk=pk)
    stock_request.delete()
    messages.success(request, "Stock Request deleted successfully.")
    return redirect('stock_request_list')

def bulk_upload_stock_request(request):
    if request.method == 'POST' and request.FILES.get('excel_file'):
        file = request.FILES['excel_file']
        errors = []

        try:
            wb = openpyxl.load_workbook(file)
            ws = wb.active

            # Get headers
            headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

            # Required columns
            required_cols = ['warehouse_name', 'product_sku', 'quantity_requested']
            for col in required_cols:
                if col not in headers:
                    messages.error(request, f"Missing required column: {col}")
                    return redirect('bulk_upload_stock_request')

            # Create list of dicts for each row
            rows = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                row_dict = dict(zip(headers, row))
                rows.append(row_dict)

        except Exception as e:
            messages.error(request, f"Error reading Excel file: {str(e)}")
            return redirect('bulk_upload_stock_request')

        # Group by warehouse_name
        from collections import defaultdict
        grouped_data = defaultdict(list)
        for row in rows:
            warehouse_name = str(row.get('warehouse_name')).strip()
            grouped_data[warehouse_name].append(row)

        # Process each group
        for warehouse_name, group_rows in grouped_data.items():
            try:
                warehouse = Warehouse.objects.get(name=warehouse_name)
            except Warehouse.DoesNotExist:
                errors.append(f"Warehouse not found: {warehouse_name}")
                continue

            remarks = group_rows[0].get('remarks', '')
            request_obj = StockRequest.objects.create(
                requesting_warehouse=warehouse,
                requested_by=request.user,
                remarks=remarks or ''
            )

            for row in group_rows:
                try:
                    product = EducationalProduct.objects.get(sku=row.get('product_sku'))
                    qty = int(row.get('quantity_requested', 0))
                    StockRequestItem.objects.create(
                        stock_request=request_obj,
                        product=product,
                        quantity_requested=qty
                    )
                except Exception as e:
                    errors.append(f"Error for SKU {row.get('product_sku')}: {str(e)}")

        if errors:
            messages.warning(request, f"Some errors occurred: {errors}")
        else:
            messages.success(request, "Stock Requests uploaded successfully.")

        return redirect('stock_request_list')

    return render(request, 'inventory/bulk_upload_stock_request.html')

def download_sample_stock_request_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Sample Stock Request"
    ws.append(['warehouse_name', 'product_sku', 'quantity_requested', 'remarks'])
    ws.append(['Hyderabad WH', '4000001234', 50, 'Urgent'])
    ws.append(['Hyderabad WH', '4000005678', 30, ''])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=sample_stock_request.xlsx'
    wb.save(response)
    return response


def stock_request_list(request):
    requests = StockRequest.objects.all().order_by('-id')
    warehouses = Warehouse.objects.all()
    return render(request, 'inventory/stock_request_list.html', { 'requests': requests,'warehouses': warehouses,'page_title': 'All Stock Requests' })


def is_warehouse_head(user):
    return user.is_authenticated and user.role == 'WarehouseHead'

@login_required
@user_passes_test(is_warehouse_head)
def stock_request_action(request):
    if request.method == 'POST':
        request_id = request.POST.get('stock_request_id')
        action_type = request.POST.get('action_type')
        from_warehouse_id = request.POST.get('from_warehouse')

        stock_request = get_object_or_404(StockRequest, pk=request_id)

        if stock_request.status != 'PENDING':
            messages.warning(request, "This request is already processed.")
            return redirect('stock_request_list')

        if action_type == 'approve':
            stock_request.status = 'APPROVED'
            stock_request.approved_by = request.user
            stock_request.approved_at = timezone.now()
            stock_request.save()

            if from_warehouse_id:
                from_warehouse = get_object_or_404(Warehouse, pk=from_warehouse_id)

                for item in stock_request.items.all():
                    StockTransfer.objects.create(
                        product=item.product,
                        from_warehouse=from_warehouse,
                        to_warehouse=stock_request.requesting_warehouse,
                        quantity=item.quantity_approved or item.quantity_requested,
                        requested_by=stock_request.requested_by,
                        approved_by=request.user,
                        status='PENDING',
                        reason=f"Auto-generated from Stock Request #{stock_request.id}"
                    )

                messages.success(request, f"Stock request #{stock_request.id} approved and transfer created.")
            else:
                messages.success(request, f"Stock request #{stock_request.id} approved (no transfer created).")

        elif action_type == 'reject':
            stock_request.status = 'REJECTED'
            stock_request.approved_by = request.user
            stock_request.approved_at = timezone.now()
            stock_request.save()
            messages.success(request, f"Stock request #{stock_request.id} rejected.")

        else:
            messages.error(request, "Invalid action.")
            return redirect('stock_request_list')

        return redirect('stock_request_list')


def export_stock_requests_excel(request, request_id):
    stock_request = StockRequest.objects.prefetch_related('items').get(id=request_id)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Stock Request #{stock_request.id}"

    headers = ["Request ID", "Warehouse", "Requested By", "Status", "Product", "Qty", "Remarks"]
    ws.append(headers)

    for item in stock_request.items.all():
        ws.append([
            stock_request.id,
            stock_request.requesting_warehouse.name,
            stock_request.requested_by.username if stock_request.requested_by else "N/A",
            stock_request.status,
            item.product.product_description,
            item.quantity_requested,
            stock_request.remarks
        ])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=stock_request_{stock_request.id}.xlsx'
    wb.save(response)
    return response


# @login_required
# def stock_transfer_list(request):
#     # Optionally filter based on user's warehouse
#     user = request.user

#     if hasattr(user, 'warehouse'):  # assuming this relation exists
#         transfers = StockTransfer.objects.filter(
#             from_warehouse=user.warehouse
#         ) | StockTransfer.objects.filter(
#             to_warehouse=user.warehouse
#         )
#     else:
#         transfers = StockTransfer.objects.all()

#     return render(request, 'inventory/stock_transfer_list.html', {
#         'transfers': transfers.order_by('-requested_at'),
#         'page_title': 'Stock Transfers'
#     })


# from django.views.decorators.http import require_POST
# from django.core.exceptions import ObjectDoesNotExist

# @login_required
# @require_POST
# def complete_stock_transfer(request, transfer_id):
#     transfer = get_object_or_404(StockTransfer, id=transfer_id)
#     if transfer.status != 'IN_TRANSIT':
#         messages.warning(request, "Only transfers in transit can be marked as completed.")
#         return redirect('stock_transfer_list')

#     try:
#         transfer.status = 'COMPLETED'
#         transfer.completed_at = timezone.now()
#         transfer.save()
#         messages.success(request, f"Transfer #{transfer.id} marked as COMPLETED.")
#     except Exception as e:
#         messages.error(request, f"Error: {e}")
#     return redirect('stock_transfer_list')


# @login_required
# @require_POST
# def mark_transfer_in_transit(request, transfer_id):
#     transfer = get_object_or_404(StockTransfer, id=transfer_id)
#     if transfer.status != 'PENDING':
#         messages.warning(request, "Only pending transfers can be moved to transit.")
#         return redirect('stock_transfer_list')

#     try:
#         transfer.status = 'IN_TRANSIT'
#         transfer.approved_by = request.user
#         transfer.save()
#         messages.success(request, f"Transfer #{transfer.id} marked as IN TRANSIT.")
#     except Exception as e:
#         messages.error(request, f"Error: {e}")
#     return redirect('stock_transfer_list')


@login_required
def stock_transfer_list(request):
    transfers = StockTransfer.objects.select_related('product', 'from_warehouse', 'to_warehouse').order_by('-requested_at')
    return render(request, 'inventory/stock_transfer_list.html', {
        'transfers': transfers,
        'page_title': 'Stock Transfers'
    })


@login_required
def update_stock_transfer_status(request, transfer_id, status):
    transfer = get_object_or_404(StockTransfer, pk=transfer_id)

    status = status.upper()
    if status not in ['IN_TRANSIT', 'COMPLETED', 'CANCELLED']:
        messages.error(request, "Invalid status.")
        return redirect('stock_transfer_list')

    try:
        if status == 'IN_TRANSIT':
            # Check from_warehouse stock
            try:
                from_item = InventoryItem.objects.get(
                    product=transfer.product,
                    warehouse=transfer.from_warehouse
                )
            except InventoryItem.DoesNotExist:
                messages.error(request, f"Inventory record not found for {transfer.from_warehouse.name}")
                return redirect('stock_transfer_list')

            if from_item.quantity_in_stock < transfer.quantity:
                messages.error(request, f"Not enough stock in {transfer.from_warehouse.name}")
                return redirect('stock_transfer_list')

            from_item.quantity_in_stock -= transfer.quantity
            from_item.save()

            # Log transfer out
            StockHistory.objects.create(
                product=transfer.product,
                warehouse=transfer.from_warehouse,
                previous_quantity=from_item.quantity_in_stock + transfer.quantity,
                changed_quantity=-transfer.quantity,
                new_quantity=from_item.quantity_in_stock,
                action_type='TRANSFER_OUT',
                reference_id=f"TRANSFER-{transfer.id}",
                changed_by=request.user
            )

        elif status == 'COMPLETED':
            to_item, _ = InventoryItem.objects.get_or_create(
                product=transfer.product,
                warehouse=transfer.to_warehouse
            )
            to_item.quantity_in_stock += transfer.quantity
            to_item.save()

            # Log transfer in
            StockHistory.objects.create(
                product=transfer.product,
                warehouse=transfer.to_warehouse,
                previous_quantity=to_item.quantity_in_stock - transfer.quantity,
                changed_quantity=transfer.quantity,
                new_quantity=to_item.quantity_in_stock,
                action_type='TRANSFER_IN',
                reference_id=f"TRANSFER-{transfer.id}",
                changed_by=request.user
            )

            transfer.completed_at = timezone.now()

        transfer.status = status
        transfer.save()

        messages.success(request, f"Stock transfer marked as {status}.")
    except Exception as e:
        messages.error(request, f"Error: {str(e)}")

    return redirect('stock_transfer_list')