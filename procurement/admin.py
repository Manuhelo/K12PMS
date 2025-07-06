from django.contrib import admin
from django import forms
from django.utils.html import format_html
from django.shortcuts import redirect, render
from django.urls import path, reverse
from django.contrib import messages
from django.template.response import TemplateResponse
import csv
import openpyxl
from openpyxl import load_workbook
import io
from .forms import VendorQuotationUploadForm
from .models import RFQ, RFQItem, VendorQuotation, VendorBid, PurchaseOrder, DeliveryTracking
from purchase_requests.models import RequestItem, PurchaseRequest
from vendors.models import Vendor
import openpyxl
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
import uuid
from babel.numbers import format_currency
from django.utils import timezone
from .utils import generate_po_barcode
from django.contrib import messages


@admin.action(description="Generate barcode for selected PO(s)")
def generate_barcode_for_po(modeladmin, request, queryset):
    for po in queryset:
        path = generate_po_barcode(po.po_number)
        messages.success(request, f"Barcode generated for PO: {po.po_number} at {path}")

def generate_custom_po_number():
    prefix = "24"
    zeros = "0000"

    # Get last PO's sequence number
    last_po = PurchaseOrder.objects.order_by('-id').first()
    if last_po and last_po.po_number:
        # last_po.po_number looks like "240000xxxx"
        # extract last 4 digits and convert to int
        last_seq_str = last_po.po_number[-4:]
        try:
            last_seq = int(last_seq_str)
        except ValueError:
            last_seq = 0
        new_seq = last_seq + 1
    else:
        new_seq = 1  # start from 0001 if none exists

    # pad sequence to 4 digits
    seq_str = f"{new_seq:04d}"

    po_number = f"{prefix}{zeros}{seq_str}"
    return po_number


admin.site.register(RFQItem)

# Inline for RFQ Items
class RFQItemInline(admin.TabularInline):
    model = RFQItem
    extra = 0
    readonly_fields = ['request_item',]

@admin.register(RFQ)
class RFQAdmin(admin.ModelAdmin):
    list_display = ('id', 'purchase_request', 'created_by', 'created_at', 'status')
    inlines = [RFQItemInline]


class VendorQuotationInline(admin.TabularInline):
    model = VendorQuotation
    extra = 0
    readonly_fields = ['rfq_item','quantity','quoted_price', 'total_price','lead_time_days','remarks']

    def quantity(self, obj):
        try:
            return obj.rfq_item.request_item.quantity
        except AttributeError:
            return "-"
    quantity.short_description = "Requested Qty"

    def total_price(self, obj):
        try:
            qty = obj.rfq_item.request_item.quantity
            return qty * obj.quoted_price
        except (AttributeError, TypeError):
            return "-"
    total_price.short_description = "Total Price"

@admin.register(VendorBid)
class VendorBidAdmin(admin.ModelAdmin):
    list_display = ('rfq','purchase_request_description', 'vendor', 'submission_group','status','total_items', 'total_cost','approve_button', 'reject_button', 'submitted_at')  # Show these columns in admin list view
    #list_filter = ('status', 'rfq',)  # Add filter options in the sidebar
    search_fields = ('vendor__name',)  # Enable search on related fields
    readonly_fields = ('submitted_at',)  # Make 'submitted_at' read-only in form
    inlines = [VendorQuotationInline]
    change_list_template = "admin/vendorquotation_change_list.html"
    actions = ['export_as_excel']

    def purchase_request_description(self, obj):
        return obj.rfq.purchase_request.description
    purchase_request_description.short_description = 'PR Description'

    def total_items(self, obj):
        return obj.quotations.count()
    total_items.short_description = "Total Items"

    def total_cost(self, obj):
        quotations = obj.quotations.select_related('rfq_item__request_item')
        total = 0
        for q in quotations:
            quantity = q.rfq_item.request_item.quantity if q.rfq_item and q.rfq_item.request_item else 1
            if q.quoted_price is not None:
                total += q.quoted_price * quantity
        return format_currency(total, 'INR', locale='en_IN')
    total_cost.short_description = "Total Cost"
    
    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path(
            "<int:pk>/approve/",
            self.admin_site.admin_view(self.approve_quotation),
            name="vendorquotation_approve",  # optional name for reverse()
        ),
        path("<int:pk>/reject/", self.admin_site.admin_view(self.reject_quotation), name="vendorquotation_reject"),
        path(
                "upload-csv/",
                self.admin_site.admin_view(self.upload_csv),
                name="vendorquotation_upload_csv",  # important!
            ),
        path('export/<int:pk>/', self.admin_site.admin_view(self.export_excel), name='vendorbid_export'),   
        ]
        return custom_urls + urls
    
    def change_view(self, request, object_id, form_url='', extra_context=None):
        print("CHANGE VIEW CALLED FOR VendorBid")  # Debug line
        extra_context = extra_context or {}
        export_url = reverse('admin:vendorbid_export', args=[object_id])
        extra_context['export_url'] = export_url
        return super().change_view(request, object_id, form_url, extra_context=extra_context)

    def approve_button(self, obj):
        if obj.status == 'Approved':
            return "✅ Approved"
        elif obj.status == 'Rejected':
            return "-"
        url = reverse("admin:vendorquotation_approve", args=[obj.pk])
        return format_html('<a class="button" href="{}" style="color:green;">Approve</a>', url)

    approve_button.short_description = "Approve"

    def reject_button(self, obj):
        if obj.status == 'Rejected':
            return "❌ Rejected"
        elif obj.status == 'Approved':
            return "-"
        url = reverse("admin:vendorquotation_reject", args=[obj.pk])
        return format_html('<a class="button" href="{}" style="color:red;">Reject</a>', url)

    reject_button.short_description = "Reject"


    def approve_quotation(self, request, pk):
        try:
            quotation = VendorBid.objects.get(pk=pk)
            rfq = quotation.rfq
            vendor = quotation.vendor

            # Reject all other quotations for this RFQ item
            VendorBid.objects.filter(rfq=rfq,vendor=vendor).exclude(pk=quotation.pk).update(status='Rejected')

            # Approve this one
            quotation.status = 'Approved'
            quotation.save()

            self.message_user(request, "Vendor quotation approved. Others rejected.", level=messages.SUCCESS)
        except VendorBid.DoesNotExist:
            self.message_user(request, "VendorQuotation not found.", level=messages.ERROR)

        return redirect("../")
    
    def reject_quotation(self, request, pk):
        try:
            quotation = VendorBid.objects.get(pk=pk)
            quotation.status = 'Rejected'
            quotation.save()

            self.message_user(request, "Vendor quotation rejected.", level=messages.SUCCESS)
        except VendorBid.DoesNotExist:
            self.message_user(request, "Vendor quotation not found.", level=messages.ERROR)

        return redirect("../")
    
    def upload_csv(self, request):
        upload_stats = None  # Initialize it at the top
        if request.method == "POST":
            form = VendorQuotationUploadForm(request.POST, request.FILES)
            if form.is_valid():
                upload_file = form.cleaned_data['csv_file']
                file_name = upload_file.name
                # After confirming the file is valid, generate a new submission_group for this CSV upload:
                submission_group = uuid.uuid4()
        
                unmatched_rows = []  # For rows with SKUs not found

                try:
                    if file_name.endswith(".csv"):
                        data = upload_file.read().decode('utf-8')
                        io_string = io.StringIO(data)
                        reader = csv.DictReader(io_string)
                    elif file_name.endswith(".xlsx") or file_name.endswith(".xls"):
                        wb = load_workbook(upload_file, data_only=True)
                        sheet = wb.active
                        headers = [str(cell.value).strip().lower() for cell in sheet[1]]
                        rows = []
                        for row in sheet.iter_rows(min_row=2, values_only=True):
                            row_dict = dict(zip(headers, row))
                            rows.append(row_dict)

                    else:
                        self.message_user(request, "Unsupported file format. Upload CSV or Excel.", level=messages.ERROR)
                        return redirect("..")
                    
                    upload_stats = {}

                    for row in rows:
                        try:
                            purchase_request_number = str(row.get('purchase_request_number')).strip()
                            sku = str(row.get('sku')).strip().upper()
                            vendor_value = str(row.get('vendor') or row.get('vendor_id') or row.get('vendor_name')).strip()
                            quoted_price = row.get('quoted_price')
                            lead_time_days = row.get('lead_time_days')
                            remarks = row.get('remarks', '')

                        # Validate PurchaseRequest
                            try:
                                pr = PurchaseRequest.objects.get(request_number=purchase_request_number)
                            except PurchaseRequest.DoesNotExist:
                                unmatched_rows.append({**row, 'error': 'PurchaseRequest not found'})
                                continue

                            request_items = RequestItem.objects.filter(purchase_request=pr)
                            total_items = request_items.count()

                            if purchase_request_number not in upload_stats:
                                upload_stats[purchase_request_number] = {"uploaded": 0, "total": total_items, "total_rows_in_file": 0}

                            # Validate SKU in RequestItem via EducationalProduct
                            try:
                                request_item = RequestItem.objects.get(
                                    purchase_request=pr,
                                    product__sku=sku
                                )
                            except RequestItem.DoesNotExist:
                                unmatched_rows.append({**row, 'error': 'SKU not found in RequestItem'})
                                continue
                            
                            # Get or create RFQ
                            rfq, _ = RFQ.objects.get_or_create(
                                purchase_request=pr,
                                defaults={'created_by': request.user}
                            )

                            # Create or get RFQItem
                            rfq_item, _ = RFQItem.objects.get_or_create(
                            rfq=rfq,
                            request_item=request_item
                            )

                            # Resolve vendor (by name or ID)
                            try:
                                if vendor_value.isdigit():
                                    vendor = Vendor.objects.get(id=int(vendor_value))
                                else:
                                    vendor = Vendor.objects.get(name__iexact=vendor_value.strip())
                            except Vendor.DoesNotExist:
                                unmatched_rows.append({**row, 'error': 'Vendor not found'})
                                continue

                       

                            # Get VendorBid for this vendor and RFQ
                            vendor_bid, _ = VendorBid.objects.get_or_create(
                                rfq=rfq,
                                vendor=vendor,
                                submission_group=submission_group,
                                defaults={'status': 'Pending'}
                            )

                            # Create VendorQuotation
                            VendorQuotation.objects.update_or_create(
                                rfq_item=rfq_item,
                                vendor_bid=vendor_bid,
                                defaults={
                                    'quoted_price': quoted_price,
                                    'lead_time_days': lead_time_days,
                                    'remarks': remarks,
                                }
                            )
                            upload_stats[purchase_request_number]["uploaded"] += 1  # Increment successful uploads
                            upload_stats[purchase_request_number]["total_rows_in_file"] += 1

                        except Exception as e:
                            unmatched_rows.append({**row, 'error': str(e)})

                # If errors occurred, generate Excel file
                    if unmatched_rows:
                        response = HttpResponse(
                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                        )
                        response['Content-Disposition'] = 'attachment; filename=not_found_skus.xlsx'

                        wb = load_workbook(io.BytesIO(), write_only=True)
                        ws = wb.create_sheet("Unmatched Rows")
                        headers = list(unmatched_rows[0].keys())
                        ws.append(headers)

                        for row in unmatched_rows:
                            ws.append([row.get(col, '') for col in headers])

                        # Autofit column widths
                        for col in ws.columns:
                            max_length = max(len(str(cell.value or "")) for cell in col)
                            ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2

                        wb.save(response)
                        return response

                    self.message_user(request, "Vendor quotations uploaded successfully.", level=messages.SUCCESS)
                    #return redirect("..")
                    form = VendorQuotationUploadForm()  # Reset form after upload
                    context = {
                        'form': form,
                        'title': 'Upload Vendor Quotation CSV',
                        'upload_stats': upload_stats,
                        'unmatched_rows': unmatched_rows,  # Add this
                    }
                    return render(request, "admin/upload_csv_form.html", context)

                except Exception as e:
                    self.message_user(request, f"Error processing file: {str(e)}", level=messages.ERROR)
                    return redirect("..")

        else:
            form = VendorQuotationUploadForm()
            context = {
                'form': form,
                'title': 'Upload Vendor Quotation CSV',
                'upload_stats': upload_stats  # Add this line
            }
        return render(request, "admin/upload_csv_form.html", context)
    
    def export_excel(self, request, pk):

        try:
            bid = VendorBid.objects.get(pk=pk)
        except VendorBid.DoesNotExist:
            self.message_user(request, "VendorBid not found.", level=messages.ERROR)
            return redirect("..")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Vendor Bids"

        headers = [
            "Vendor", "RFQ", "Purchase Request", "Item Name", "SKU", "Quantity",
            "Quoted Price","Total Price", "Lead Time (days)", "Remarks", "Submission Group", "Status", 
        ]
        ws.append(headers)

        for quotation in bid.quotations.select_related('rfq_item__request_item__product'):
            item = quotation.rfq_item.request_item
            product = item.product
            quantity = item.quantity or 1
            quoted_price = quotation.quoted_price or 0
            total_price = quoted_price * quantity

            ws.append([
                bid.vendor.name,
                str(bid.rfq.id),
                item.purchase_request.request_number,
                product.product_description,
                product.sku,
                quantity,
                quoted_price,
                total_price,
                quotation.lead_time_days,
                quotation.remarks,
                str(bid.submission_group),
                bid.status,
                
            ])

        # Autofit column widths
        for column_cells in ws.columns:
            length = max(len(str(cell.value or "")) for cell in column_cells)
            ws.column_dimensions[get_column_letter(column_cells[0].column)].width = length + 2

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="vendor_bids.xlsx"'
        wb.save(response)
        return response
    

    def save_model(self, request, obj, form, change):
        was_approved = False
        if change:
            old_obj = VendorBid.objects.get(pk=obj.pk)
            # Check if status changed to Approved
            if old_obj.status != 'Approved' and obj.status == 'Approved':
                was_approved = True
        
        super().save_model(request, obj, form, change)

        if was_approved:
            # Create PO if not exists for this VendorBid
            po, created = PurchaseOrder.objects.get_or_create(vendor_bid=obj)
            if created:
                # Generate your custom PO number here
                po.po_number = generate_custom_po_number()
                po.issued_by = request.user
                po.save()

    

# admin.site.register(VendorBid,VendorBidAdmin)

@admin.register(VendorQuotation)
class VendorQuotationAdmin(admin.ModelAdmin):
    list_display = ('get_vendor', 'rfq_item', 'quoted_price', 'lead_time_days', 'get_submitted_at')
    

    def get_vendor(self, obj):
        if obj.vendor_bid and obj.vendor_bid.vendor:
            return obj.vendor_bid.vendor.name  # or however you want to display the vendor
        return "-"

    def get_submitted_at(self, obj):
        if obj.vendor_bid and obj.vendor_bid.submitted_at:
            return obj.vendor_bid.submitted_at
        return "-"


from django.utils.html import format_html
from django.urls import reverse

@admin.register(PurchaseOrder)
class PurchaseOrderAdmin(admin.ModelAdmin):
    list_display = ('po_number', 'vendor_bid', 'status', 'issued_at','download_barcode')
    #actions = [generate_barcode_for_po]

    def download_barcode(self, obj):
        return format_html('<a href="/procurement/po/{}/barcode/" target="_blank">Download Barcode</a>', obj.id)

    download_barcode.short_description = 'Download Barcode'
